
from sys import stdout
from turtle import width
import openpyxl
from ..models import Gorev_listesi,Kalibrator_listesi
import datetime
from django.contrib.auth.models import User
from django.shortcuts import render, redirect


def excelLoad(files):
    excel_data = model_write(files)

    for excel_cell in excel_data[1:]:
        # datefile  excelden gelen 2032-05-26 00:00:00 formatı kabul etmedi
        # datefile(YYYY-MM-DD) formatına uyarlamak için
        if "Biyomedikal Kalibrasyon Laboratuvarı" in excel_cell:
    
            period_of_validity_dd=excel_cell[7][0:2]
            period_of_validity_mm=excel_cell[7][3:5]
            period_of_validity_yy=excel_cell[7][6:10]
            period_of_validity = datetime.date(year=int(period_of_validity_yy), month=int(period_of_validity_mm), day=int(period_of_validity_dd))
            print(period_of_validity)
            Kalibrator_listesi.objects.create(inventory_code=excel_cell[0],inventory_type=excel_cell[1],inventory_sub_type=excel_cell[2],device_brand=excel_cell[3],device_model=excel_cell[4],serial_number=excel_cell[5],period_of_validity=period_of_validity,traceability=excel_cell[8])
            

        else:
            if excel_cell[9]!="" :
                plan_date_DD = excel_cell[9][0:2]
                plan_date_MM = excel_cell[9][3:5]
                plan_date_YYYY = excel_cell[9][6:10]
                if bool(excel_cell[14])==False:
                    excel_cell[14]=None
        
                plan_date = datetime.date(year=int(plan_date_YYYY), month=int(
                    plan_date_MM), day=int(plan_date_DD))

                ### username#####
                Username_get = str(excel_cell[19])
                # turkish to english
                Username_trans = Username_get.maketrans(
                    "ğĞıİöÖüÜşŞçÇ ", "gGiIoÖüÜşŞcC.")
                Username = Username_get.translate(Username_trans).lower()
                
                user_id = User.objects.get(username=Username).pk

                Gorev_listesi.objects.create(task_number=excel_cell[0], inventory_code=excel_cell[1], inventory_type=excel_cell[2], inventory_sub_type=excel_cell[3], device_brand=excel_cell[4], device_model=excel_cell[5], serial_number=excel_cell[6], task_type=excel_cell[7], task_name=excel_cell[8], plan_date=plan_date, task_status=excel_cell[10],
                                            explanation=excel_cell[11], liable_type=excel_cell[12], responsible=excel_cell[13], completion_date=excel_cell[14], branch=excel_cell[15], branch_building=excel_cell[16], building_floor=excel_cell[17], location=excel_cell[18], person_doing=excel_cell[19], label=excel_cell[20], user_id=str(user_id))
            
            else:
                raise Exception("Plan Tarihini kontrol ediniz boş data var")

def model_write(files):
    excel_file = openpyxl.load_workbook(files)
    sheet_name = excel_file.sheetnames[0]
    sheet = excel_file[sheet_name]
    excel_data = []
    for row in sheet.iter_rows():
        row_data = []

        for cell in row:

            row_data.append(str(cell.value).replace("None", ""))
        excel_data.append(row_data)
    return excel_data
